import os
import re
import warnings
from collections import Counter

import pandas as pd
import openpyxl

from fastapi import FastAPI, UploadFile, File, Body
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

from openpyxl.styles import PatternFill

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.pipeline import Pipeline


warnings.filterwarnings("ignore", category=UserWarning)

app = FastAPI()

os.makedirs("uploads", exist_ok=True)
os.makedirs("outputs", exist_ok=True)


# -----------------------------
# TEXT CLEANING
# -----------------------------
def clean_text(text):
    text = str(text).lower().strip()
    text = re.sub(r"[^a-z0-9 ]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def tokenize(text):
    text = clean_text(text)
    if not text:
        return []
    return text.split()


# -----------------------------
# LOAD TRAINING DATA
#
# training.xlsx:
#   skip first row
#   col 0 = Account Item
#   col 1 = Description / Keyword
#   col 2 = GST sign
#
# feedback.csv (optional):
#   same 3 columns
# -----------------------------
def load_data():
    df = pd.read_excel(
        "training.xlsx",
        skiprows=1,
        header=None,
        engine="openpyxl"
    )

    if os.path.exists("feedback.csv"):
        fb = pd.read_csv("feedback.csv", header=None)
        df = pd.concat([df, fb], ignore_index=True)

    if df.shape[1] < 3:
        raise ValueError(
            "training.xlsx must have at least 3 columns: "
            "Account Item, Description/Keyword, GST sign"
        )

    df = df.dropna(how="all")

    out = pd.DataFrame()
    out["Code1"] = df[0].fillna("").astype(str).str.strip()
    out["Keyword"] = df[1].fillna("").astype(str).apply(clean_text)
    out["Code2"] = df[2].fillna("").astype(str).str.strip()

    out = out[out["Code1"] != ""]
    out = out[out["Keyword"] != ""]

    # Remove accidental header-like rows
    out = out[
        ~(
            out["Code1"].str.lower().isin(
                ["code1", "account item", "accountitem", "predictedcode1"]
            )
            |
            out["Keyword"].str.lower().isin(
                ["keyword", "description", "desc"]
            )
            |
            out["Code2"].str.lower().isin(
                ["code2", "gst sign", "gstsign", "predictedcode2"]
            )
        )
    ].copy()

    out = out.reset_index(drop=True)

    print("✅ Training rows used:", len(out))
    print("✅ Sample descriptions:", out["Keyword"].head(5).tolist())
    print("✅ Sample Account Item:", out["Code1"].head(5).tolist())
    print("✅ Sample GST sign:", out["Code2"].head(5).tolist())

    if len(out) == 0:
        raise ValueError("No valid training rows found after removing blank rows.")

    return out


# -----------------------------
# BUILD DESCRIPTION-LEVEL INDEX
# Used for exact / contains / similarity on row level
# -----------------------------
def build_description_index():
    vectorizer = TfidfVectorizer(
        stop_words=None,
        ngram_range=(1, 2),
        max_features=8000,
        token_pattern=r"(?u)\b\w+\b"
    )
    matrix = vectorizer.fit_transform(train_df["Keyword"])
    return vectorizer, matrix


# -----------------------------
# BUILD ACCOUNT ITEM PROFILE MAP
#
# New upgrade:
# Each Account Item can map to many descriptions
# and many GST signs.
# -----------------------------
def build_account_item_profiles():
    grouped_rows = []

    for code1, grp in train_df.groupby("Code1", dropna=False):
        keywords = grp["Keyword"].dropna().astype(str).tolist()
        keywords = [k for k in keywords if k.strip() != ""]
        unique_keywords = list(dict.fromkeys(keywords))

        code2_values = grp["Code2"].fillna("").astype(str).tolist()
        non_blank_code2 = [x.strip() for x in code2_values if str(x).strip() != ""]

        default_code2 = ""
        code2_counts = {}
        if non_blank_code2:
            cnt = Counter(non_blank_code2)
            default_code2 = cnt.most_common(1)[0][0]
            code2_counts = dict(cnt)

        grouped_rows.append({
            "Code1": str(code1),
            "CombinedText": " || ".join(unique_keywords),
            "Descriptions": unique_keywords,
            "DefaultCode2": default_code2,
            "Code2Counts": code2_counts,
            "DescriptionCount": len(unique_keywords)
        })

    profiles_df = pd.DataFrame(grouped_rows)

    vectorizer = TfidfVectorizer(
        stop_words=None,
        ngram_range=(1, 2),
        max_features=8000,
        token_pattern=r"(?u)\b\w+\b"
    )
    matrix = vectorizer.fit_transform(profiles_df["CombinedText"])

    return profiles_df.reset_index(drop=True), vectorizer, matrix


# -----------------------------
# BUILD ACCOUNT ITEM ML MODEL
# Fallback only
# -----------------------------
def build_code1_model():
    model = Pipeline([
        ("tfidf", TfidfVectorizer(
            stop_words=None,
            ngram_range=(1, 2),
            max_features=8000,
            token_pattern=r"(?u)\b\w+\b"
        )),
        ("clf", LogisticRegression(
            max_iter=3000,
            C=3,
            class_weight="balanced"
        ))
    ])

    model.fit(train_df["Keyword"], train_df["Code1"])
    return model


# -----------------------------
# GLOBAL DATA / MODELS
# -----------------------------
train_df = load_data()
desc_vectorizer, desc_matrix = build_description_index()
profiles_df, profile_vectorizer, profile_matrix = build_account_item_profiles()
code1_model = build_code1_model()

print("✅ Engine ready: exact → contains → token overlap → account-item profile → ML fallback")


# -----------------------------
# OPTIONAL BUSINESS RULES
# Return:
#   Account Item, GST sign, confidence
# -----------------------------
def apply_rules(text):
    text = clean_text(text)

    # Example:
    # if text == "deposit":
    #     return "DepositAccount", "", 0.99

    return None, None, None


# -----------------------------
# EXACT DESCRIPTION MATCH
# -----------------------------
def exact_match(text_clean):
    rows = train_df[train_df["Keyword"] == text_clean]
    if len(rows) == 0:
        return None

    code1 = rows["Code1"].mode().iloc[0]
    code2_series = rows["Code2"].fillna("").astype(str)
    code2 = code2_series.mode().iloc[0] if len(code2_series) else ""

    return {
        "code1": str(code1),
        "code2": str(code2),
        "score": 1.00,
        "method": "exact"
    }


# -----------------------------
# CONTAINS MATCH
# -----------------------------
def contains_match(text_clean):
    rows = train_df[
        train_df["Keyword"].apply(
            lambda k: k != "" and (k in text_clean or text_clean in k)
        )
    ]

    if len(rows) == 0:
        return None

    rows = rows.copy()
    rows["kw_len"] = rows["Keyword"].astype(str).apply(len)
    rows = rows.sort_values("kw_len", ascending=False)

    best_code1 = rows["Code1"].mode().iloc[0]
    best_code2_series = rows["Code2"].fillna("").astype(str)
    best_code2 = best_code2_series.mode().iloc[0] if len(best_code2_series) else ""

    return {
        "code1": str(best_code1),
        "code2": str(best_code2),
        "score": 0.96,
        "method": "contains"
    }


# -----------------------------
# TOKEN OVERLAP MATCH
# Row-level overlap scoring
# -----------------------------
def token_overlap_match(text_clean):
    input_tokens = set(tokenize(text_clean))
    if not input_tokens:
        return None

    best_idx = None
    best_score = 0.0

    for idx, row in train_df.iterrows():
        row_tokens = set(tokenize(row["Keyword"]))
        if not row_tokens:
            continue

        overlap = len(input_tokens & row_tokens)
        if overlap == 0:
            continue

        # bias towards how much of the training keyword is matched
        score = overlap / max(1, len(row_tokens))

        # slight boost when most input is also covered
        input_cover = overlap / max(1, len(input_tokens))
        score = (0.7 * score) + (0.3 * input_cover)

        if score > best_score:
            best_score = score
            best_idx = idx

    if best_idx is None:
        return None

    if best_score < 0.45:
        return None

    return {
        "code1": str(train_df.iloc[best_idx]["Code1"]),
        "code2": str(train_df.iloc[best_idx]["Code2"]),
        "score": round(min(0.93, best_score), 2),
        "method": "token_overlap"
    }


# -----------------------------
# ACCOUNT ITEM PROFILE MATCH
#
# New upgrade:
# compare input text to the combined descriptions
# of each Account Item
# -----------------------------
def account_item_profile_match(text_clean):
    vec = profile_vectorizer.transform([text_clean])
    sims = cosine_similarity(vec, profile_matrix)[0]

    best_idx = sims.argmax()
    best_score = float(sims[best_idx])

    best_code1 = str(profiles_df.iloc[best_idx]["Code1"])
    default_code2 = str(profiles_df.iloc[best_idx]["DefaultCode2"])

    return {
        "code1": best_code1,
        "code2": default_code2,
        "score": round(best_score, 2),
        "method": "profile"
    }


# -----------------------------
# CHOOSE BEST GST SIGN
# for one Account Item using its multiple descriptions
# -----------------------------
def choose_gst_for_code1(code1, text_clean):
    subset = train_df[train_df["Code1"] == code1].copy()

    if len(subset) == 0:
        return "", 0.0

    # 1) exact within this Account Item
    exact_rows = subset[subset["Keyword"] == text_clean]
    if len(exact_rows) > 0:
        code2_values = exact_rows["Code2"].fillna("").astype(str).tolist()
        non_blank = [x for x in code2_values if x.strip() != ""]
        if non_blank:
            return Counter(non_blank).most_common(1)[0][0], 1.00
        return "", 1.00

    # 2) contains within this Account Item
    contains_rows = subset[
        subset["Keyword"].apply(
            lambda k: k != "" and (k in text_clean or text_clean in k)
        )
    ]
    if len(contains_rows) > 0:
        code2_values = contains_rows["Code2"].fillna("").astype(str).tolist()
        non_blank = [x for x in code2_values if x.strip() != ""]
        if non_blank:
            return Counter(non_blank).most_common(1)[0][0], 0.96
        return "", 0.96

    # 3) similarity within this Account Item
    input_vec = desc_vectorizer.transform([text_clean])
    subset_idx = subset.index.tolist()
    subset_matrix = desc_matrix[subset_idx]
    sims = cosine_similarity(input_vec, subset_matrix)[0]

    best_local_pos = sims.argmax()
    best_score = float(sims[best_local_pos])
    best_row = subset.iloc[best_local_pos]

    candidate_code2 = str(best_row["Code2"]).strip()

    if candidate_code2 != "":
        return candidate_code2, round(best_score, 2)

    # 4) fallback to profile default GST sign
    profile_row = profiles_df[profiles_df["Code1"] == code1]
    if len(profile_row) > 0:
        default_code2 = str(profile_row.iloc[0]["DefaultCode2"]).strip()
        if default_code2 != "":
            return default_code2, round(best_score * 0.9, 2)

    return "", round(best_score, 2)


# -----------------------------
# FINAL PREDICTION
#
# Flow:
#   0. rules
#   1. exact description match
#   2. contains match
#   3. token overlap match
#   4. account-item profile match
#   5. ML fallback if profile is weak
# -----------------------------
def predict_codes(text):
    text_clean = clean_text(text)

    if text_clean == "":
        return "", "", 0.0, 0.0, 0.0

    # 0. Rule layer
    rule_code1, rule_code2, rule_conf = apply_rules(text_clean)
    if rule_code1 is not None:
        return (
            str(rule_code1),
            str(rule_code2),
            float(rule_conf),
            1.00,
            1.00
        )

    # 1. exact
    exact_res = exact_match(text_clean)
    if exact_res is not None:
        return (
            exact_res["code1"],
            exact_res["code2"],
            1.00,
            1.00,
            1.00
        )

    # 2. contains
    contains_res = contains_match(text_clean)
    if contains_res is not None:
        gst_code2, gst_conf = choose_gst_for_code1(contains_res["code1"], text_clean)

        overall = round((0.7 * contains_res["score"]) + (0.3 * max(gst_conf, contains_res["score"])), 2)
        return (
            contains_res["code1"],
            gst_code2,
            overall,
            round(contains_res["score"], 2),
            round(max(gst_conf, contains_res["score"]), 2)
        )

    # 3. token overlap
    overlap_res = token_overlap_match(text_clean)
    if overlap_res is not None:
        gst_code2, gst_conf = choose_gst_for_code1(overlap_res["code1"], text_clean)

        overall = round((0.7 * overlap_res["score"]) + (0.3 * max(gst_conf, overlap_res["score"])), 2)
        return (
            overlap_res["code1"],
            gst_code2,
            overall,
            round(overlap_res["score"], 2),
            round(max(gst_conf, overlap_res["score"]), 2)
        )

    # 4. profile match
    profile_res = account_item_profile_match(text_clean)
    profile_code1 = profile_res["code1"]
    profile_score = float(profile_res["score"])

    gst_code2, gst_conf = choose_gst_for_code1(profile_code1, text_clean)

    # 5. ML fallback
    ml_code1 = str(code1_model.predict([text_clean])[0])
    ml_probs = code1_model.predict_proba([text_clean])[0]
    ml_conf = float(max(ml_probs))

    # If profile is decent, keep profile result
    if profile_score >= 0.55:
        final_code1 = profile_code1
        final_code2 = gst_code2

        code1_conf = round(profile_score, 2)
        code2_conf = round(max(gst_conf, min(0.90, profile_score)), 2)
        overall_conf = round((0.75 * code1_conf) + (0.25 * code2_conf), 2)

        return final_code1, final_code2, overall_conf, code1_conf, code2_conf

    # If profile and ML agree, boost confidence
    if ml_code1 == profile_code1:
        final_code1 = profile_code1
        final_code2 = gst_code2

        code1_conf = round(max(profile_score, ml_conf), 2)
        code2_conf = round(max(gst_conf, profile_score), 2)
        overall_conf = round((0.75 * code1_conf) + (0.25 * code2_conf), 2)

        return final_code1, final_code2, overall_conf, code1_conf, code2_conf

    # Otherwise use ML for Code1, and map GST sign from that Code1
    final_code1 = ml_code1
    final_code2, gst_conf_from_ml_code1 = choose_gst_for_code1(final_code1, text_clean)

    code1_conf = round(ml_conf, 2)
    code2_conf = round(gst_conf_from_ml_code1, 2)
    overall_conf = round((0.75 * code1_conf) + (0.25 * code2_conf), 2)

    return final_code1, final_code2, overall_conf, code1_conf, code2_conf


# -----------------------------
# REBUILD MODELS
# -----------------------------
def rebuild_models():
    global train_df
    global desc_vectorizer, desc_matrix
    global profiles_df, profile_vectorizer, profile_matrix
    global code1_model

    train_df = load_data()
    desc_vectorizer, desc_matrix = build_description_index()
    profiles_df, profile_vectorizer, profile_matrix = build_account_item_profiles()
    code1_model = build_code1_model()

    print("✅ Models rebuilt")
    print("✅ Account Item profiles:", len(profiles_df))


# -----------------------------
# UPLOAD API
#
# Input Excel:
#   start reading from row 3
#   use 3rd column as description
# -----------------------------
@app.post("/upload")
async def upload(file: UploadFile = File(...)):
    fname = file.filename.replace(" ", "_")

    input_path = f"uploads/{fname}"
    output_path = f"outputs/output_{fname}"

    with open(input_path, "wb") as f:
        f.write(await file.read())

    df = pd.read_excel(
        input_path,
        skiprows=2,
        header=None,
        engine="openpyxl"
    )

    keyword_col = 2

    if df.shape[1] <= keyword_col:
        return {"error": "Uploaded Excel does not have a 3rd column."}

    df[keyword_col] = df[keyword_col].fillna("").astype(str)

    predicted_code1 = []
    predicted_code2 = []
    overall_confidences = []
    code1_confidences = []
    code2_confidences = []

    for text in df[keyword_col]:
        code1, code2, overall_conf, code1_conf, code2_conf = predict_codes(text)

        predicted_code1.append(code1)
        predicted_code2.append(code2)
        overall_confidences.append(overall_conf)
        code1_confidences.append(code1_conf)
        code2_confidences.append(code2_conf)

    # Keep backend field names expected by locked index.html
    df["PredictedCode1"] = predicted_code1
    df["PredictedCode2"] = predicted_code2
    df["Confidence"] = overall_confidences

    df.to_excel(output_path, index=False, header=False, engine="openpyxl")

    # Highlight confidence < 80% in yellow
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    yellow = PatternFill(
        start_color="FFFF99",
        end_color="FFFF99",
        fill_type="solid"
    )

    for row in ws.iter_rows(min_row=1):
        confidence = row[-1].value
        if confidence is not None and isinstance(confidence, (int, float)) and confidence < 0.8:
            for cell in row:
                cell.fill = yellow

    wb.save(output_path)

    result_data = []
    for i, row in df.iterrows():
        result_data.append({
            "Description": row[keyword_col],
            "PredictedCode1": row["PredictedCode1"],
            "PredictedCode2": row["PredictedCode2"],
            "Confidence": row["Confidence"],
            "Code1Confidence": code1_confidences[i],
            "Code2Match": code2_confidences[i]
        })

    return {
        "download_url": f"/download/output_{fname}",
        "data": result_data
    }


# -----------------------------
# DOWNLOAD API
# -----------------------------
@app.get("/download/{file_name}")
def download(file_name: str):
    return FileResponse(
        f"outputs/{file_name}",
        filename=file_name
    )


# -----------------------------
# FEEDBACK API
#
# Saves:
#   Account Item, Description, GST sign
# -----------------------------
@app.post("/feedback")
async def feedback(data: dict = Body(...)):
    description = clean_text(data.get("description", ""))
    code1 = str(data.get("code1", "")).strip()
    code2 = str(data.get("code2", "")).strip()

    if description == "" or code1 == "":
        return {"status": "invalid"}

    row_df = pd.DataFrame([[code1, description, code2]])

    if os.path.exists("feedback.csv"):
        row_df.to_csv("feedback.csv", mode="a", header=False, index=False)
    else:
        row_df.to_csv("feedback.csv", header=False, index=False)

    return {"status": "saved"}


# -----------------------------
# BULK APPROVE API
# -----------------------------
@app.post("/approve-bulk")
async def approve_bulk(data: dict = Body(...)):
    rows = data.get("rows", [])

    if not rows:
        return {"status": "no_rows"}

    feedback_rows = []

    for row in rows:
        description = clean_text(row.get("description", ""))
        code1 = str(row.get("code1", "")).strip()
        code2 = str(row.get("code2", "")).strip()

        if description != "" and code1 != "":
            feedback_rows.append([code1, description, code2])

    if len(feedback_rows) == 0:
        return {"status": "no_valid_rows"}

    row_df = pd.DataFrame(feedback_rows)

    if os.path.exists("feedback.csv"):
        row_df.to_csv("feedback.csv", mode="a", header=False, index=False)
    else:
        row_df.to_csv("feedback.csv", header=False, index=False)

    rebuild_models()

    return {
        "status": "approved",
        "rows_saved": len(feedback_rows)
    }


# -----------------------------
# RETRAIN API
# -----------------------------
@app.post("/retrain")
def retrain():
    rebuild_models()
    return {"status": "updated"}


# -----------------------------
# STATIC UI
# -----------------------------
app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/")
def home():
    return FileResponse("static/index.html")
