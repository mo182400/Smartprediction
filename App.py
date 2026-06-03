import os
import re
import pandas as pd

from fastapi import FastAPI, UploadFile, File, Body
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

from sklearn.pipeline import Pipeline
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.svm import LinearSVC

import openpyxl
from openpyxl.styles import PatternFill

app = FastAPI()

os.makedirs("uploads", exist_ok=True)
os.makedirs("outputs", exist_ok=True)

# -----------------------------
# TEXT CLEAN
# -----------------------------
def clean_text(text):
    text = str(text).lower()
    text = re.sub(r'[^a-z0-9 ]', '', text)
    return text

# -----------------------------
# LOAD DATA
# -----------------------------
def load_data():
    df = pd.read_csv("training.csv")

    if os.path.exists("feedback.csv"):
        fb = pd.read_csv("feedback.csv")
        df = pd.concat([df, fb], ignore_index=True)

    df = df.dropna(subset=["Code"])
    df["Description"] = df["Description"].fillna("").apply(clean_text)
    df["Code"] = df["Code"].astype(int)

    return df

train_df = load_data()

# -----------------------------
# MODEL
# -----------------------------
model = Pipeline([
    ("tfidf", TfidfVectorizer(
        stop_words='english',
        ngram_range=(1,2),
        max_features=5000
    )),
    ("clf", LinearSVC())
])

model.fit(train_df["Description"], train_df["Code"])
print("✅ Model ready")

# -----------------------------
# RULES
# -----------------------------
def apply_rules(text):
    if "tax" in text:
        return 500, 0.99
    if "salary" in text:
        return 300, 0.99
    if "refund" in text:
        return 200, 0.95
    return None, None

# -----------------------------
# PREDICT
# -----------------------------
def predict(text):
    text_clean = clean_text(text)

    code, conf = apply_rules(text_clean)
    if code:
        return code, conf

    pred = model.predict([text_clean])[0]
    return int(pred), 0.75

# -----------------------------
# UPLOAD
# -----------------------------
@app.post("/upload")
async def upload(file: UploadFile = File(...)):

    fname = file.filename.replace(" ", "_")
    input_path = f"uploads/{fname}"
    output_path = f"outputs/output_{fname}"

    with open(input_path, "wb") as f:
        f.write(await file.read())

    df = pd.read_excel(input_path)
    df.columns = df.columns.str.strip().str.lower()

    desc_col = next((c for c in df.columns if "desc" in c), df.columns[0])
    df[desc_col] = df[desc_col].fillna("")

    codes, confs = [], []

    for text in df[desc_col]:
        c, conf = predict(text)
        codes.append(c)
        confs.append(round(conf, 2))

    df["PredictedCode"] = codes
    df["Confidence"] = confs
    df["LowConfidence"] = df["Confidence"] < 0.7

    df.to_excel(output_path, index=False)

    # ✅ Highlight low-confidence rows
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    red = PatternFill(start_color="FFCCCC", fill_type="solid")

    for row in ws.iter_rows(min_row=2):
        confidence = row[-2].value  # Confidence column
        if confidence < 0.7:
            for cell in row:
                cell.fill = red

    wb.save(output_path)

    return {
        "download_url": f"/download/output_{fname}",
        "data": df.to_dict(orient="records")
    }

# -----------------------------
# DOWNLOAD
# -----------------------------
@app.get("/download/{file}")
def download(file: str):
    return FileResponse(f"outputs/{file}", filename=file)

# -----------------------------
# FEEDBACK
# -----------------------------
@app.post("/feedback")
async def feedback(data: dict = Body(...)):

    df = pd.DataFrame([[data["description"], data["code"]]],
                      columns=["Description", "Code"])

    if os.path.exists("feedback.csv"):
        df.to_csv("feedback.csv", mode="a", header=False, index=False)
    else:
        df.to_csv("feedback.csv", index=False)

    return {"status": "saved"}

# -----------------------------
# RETRAIN
# -----------------------------
@app.post("/retrain")
def retrain():
    global model

    df = load_data()
    model.fit(df["Description"], df["Code"])

    return {"status": "updated"}

# -----------------------------
# UI
# -----------------------------
app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def home():
    return FileResponse("static/index.html")
``
